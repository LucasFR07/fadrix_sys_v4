import math, tablib
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from source.api.asanaApi import AsanaAPI
from source.api.shopeeApi import Shopee
from data.repository.orderV2 import OrderRepositoryV2 as ORDER


class Shopee_AutoExcel:

    CLASS_NAME = "SHOPEE_AUTOEXCEL"
    FBF = 'FBF COMUNICACAO'
    FENIX = 'FENIX SUBLIMACAO'
    XDRI = "XDRI SUBLIMACAO"    

    def __init__(self, path:str, set_colums:list):
        self.path_file = path
        self.set_colums = set_colums


    def get_order_list(self) -> list:
        """
        Obtem uma listagem com números dos pedidos e atualiza a cor da célula na planilha
        original com base na validação de cada pedido.
        """

        try:
            # Carrega a planilha com pandas para ler os dados.
            # dtype=str garante que os números de pedido não sejam interpretados como números.
            df = pd.read_excel(self.path_file, usecols=self.set_colums, dtype=str)

            # Carrega a mesma planilha com openpyxl para poder editar estilos (cores).
            workbook = load_workbook(self.path_file)
            sheet = workbook.active

            # Define as cores de preenchimento que serão usadas.
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            # Assume que a coluna de pedidos é a primeira na lista `set_colums`.
            order_column_name = self.set_colums[0]

            # Encontra o índice da coluna na planilha para saber qual coluna colorir.
            header = [cell.value for cell in sheet[1]]
            try:
                order_col_idx = header.index(order_column_name) + 1
            except ValueError:
                print(f"Erro: Coluna '{order_column_name}' não encontrada no cabeçalho da planilha.")
                return []

            processed_orders = set()
            unique_order_numbers = []

            # Itera sobre cada linha do DataFrame do pandas.
            for index, row_data in df.iterrows():
                excel_row_num = index + 2  # pandas index (0-based) to excel row (1-based) + header.
                cell_to_color = sheet.cell(row=excel_row_num, column=order_col_idx)
                order_number = row_data[order_column_name]

                if not order_number or pd.isna(order_number):
                    continue
                
                processed_orders.add(order_number)

                # Valida o pedido via API.
                shopee_response = Shopee(__class__.FBF).get_order_detail(order_number=order_number)
                
                # --- LÓGICA DE VALIDAÇÃO ---
                match shopee_response["response"]["order_list"][0]["order_status"]:
                    case "CANCELLED" | "UNPAID" | "IN_CANCEL":
                        cell_to_color.fill = red_fill # Vermelho para inválido.

                    case "READY_TO_SHIP" | "INVOICE_PENDING":
                        cell_to_color.fill = yellow_fill # Amarelo para em aberto.

                    case "PROCESSED" | "SHIPPED" | "COMPLETED":
                        cell_to_color.fill = green_fill # Verde para válido.

                unique_order_numbers.append(order_number)

            # Salva o arquivo Excel com as alterações de cor.
            workbook.save(self.path_file)

            print(f'\n🐞 GET_ORDER_LIST ==> TOTAL PEDIDOS ÚNICOS: {len(unique_order_numbers)}\n')
            print(f'\n🐞 GET_ORDER_LIST ==> PEDIDOS: {unique_order_numbers}\n')

            # O parâmetro `only_number` é um pouco redundante, pois a lista já contém apenas os números.
            # Retornando a lista para manter a consistência.
            return unique_order_numbers

        except FileNotFoundError:
            print(f'\n❌ {__class__.CLASS_NAME} > get_order_list() ==> ERRO: Arquivo não encontrado em "{self.path_file}"\n')
            return []
        except Exception as exc:
            print(f'\n❌ {__class__.CLASS_NAME} > get_order_list() ==> EXCEPTION: {exc}\n')
            return []

